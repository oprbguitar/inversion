"""Extrae el Excel de ranking a un dataset JSON normalizado.

El Excel trae varias celdas numericas con formato de fecha, por lo que openpyxl
las devuelve como datetime. Se revierten al serial de Excel (dias desde 1899-12-30).
"""
import datetime
import json
import re
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r"C:/Users/oprbg/Downloads/Ranking_Cuentas_Ahorro_Peru_Agosto_2026.xlsx")
EPOCH = datetime.datetime(1899, 12, 30)

SKIP = {"RANKING OFICIAL", "FSD JUN-AGO 2026", "TASA REFERENCIA", "FUENTES Y NOTAS"}


def num(v):
    """Devuelve el numero real detras de una celda (revierte la conversion a fecha)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, datetime.datetime):
        return round((v - EPOCH).total_seconds() / 86400, 6)
    if isinstance(v, datetime.date):
        return float((datetime.datetime(v.year, v.month, v.day) - EPOCH).days)
    if isinstance(v, datetime.time):
        return v.hour / 24 + v.minute / 1440 + v.second / 86400
    if isinstance(v, str):
        s = v.replace(",", "").replace("S/", "").strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def txt(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip() or None
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


def rows_of(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def find_row(rows, col, needle):
    for i, r in enumerate(rows):
        if col < len(r) and isinstance(r[col], str) and needle.lower() in r[col].lower():
            return i
    return None


def parse_entity(ws):
    rows = rows_of(ws)
    title = txt(rows[0][0]) or ws.title
    name, _, product = title.partition("–")
    d = {
        "sheet": ws.title,
        "entidad": name.strip() or ws.title,
        "producto": product.strip(),
        "tramos": [],
        "campos": {},
        "fuentes": {},
    }

    # Parametros del simulador (col 0/1) — la fila 3 es el encabezado.
    hdr = find_row(rows, 0, "Parámetro del simulador")
    if hdr is not None:
        for r in rows[hdr + 1: hdr + 8]:
            k = txt(r[0])
            if not k:
                continue
            d["campos"].setdefault("_sim", {})[k] = num(r[1] if len(r) > 1 else None)
        # Tramos escalonados (cols 5,6,7)
        for r in rows[hdr + 1: hdr + 12]:
            if len(r) < 8:
                continue
            desde, trea, desc = num(r[5]), num(r[6]), txt(r[7])
            if desde is None or trea is None:
                continue
            d["tramos"].append({"desde": desde, "trea": trea, "desc": desc})

    # Tabla Campo | Información | Fuente
    hdr2 = find_row(rows, 0, "Campo")
    if hdr2 is not None:
        for r in rows[hdr2 + 1:]:
            k = txt(r[0])
            if not k or k.startswith("Advertencia"):
                continue
            d["campos"][k] = txt(r[1]) if len(r) > 1 else None
            src = txt(r[2]) if len(r) > 2 else None
            if src and src.startswith("http"):
                d["fuentes"][k] = src

    sim = d["campos"].pop("_sim", {})
    d["simulador"] = {
        "saldo_ejemplo": sim.get("Saldo inicial editable (S/)"),
        "trea_max": sim.get("TREA máxima publicada"),
        "plazo_dias": sim.get("Plazo de simulación (días)"),
        "tope_remunerado": sim.get("Tope remunerado (S/)"),
        "umbral_tasa_max": sim.get("Umbral para tasa máxima (S/)"),
    }
    return d


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # --- Ranking oficial (tabla transpuesta) ---
    ws = wb["RANKING OFICIAL"]
    rows = rows_of(ws)
    hdr = find_row(rows, 0, "Entidad")
    labels = {txt(r[0]): r for r in rows[hdr:] if txt(r[0])}
    n = len(labels["Entidad"])

    def col(label, i, as_num=False):
        r = labels.get(label)
        if not r or i >= len(r):
            return None
        return num(r[i]) if as_num else txt(r[i])

    ranking = []
    for i in range(1, n):
        if not col("Entidad", i):
            continue
        ranking.append({
            "entidad": col("Entidad", i),
            "producto": col("Producto", i),
            "apertura": col("Modo de apertura", i),
            "trea": col("TREA verificada al 06/08/2026", i, True),
            "trea_video": col("TREA mostrada en el video", i, True),
            "vigencia": col("Vigencia / campaña", i),
            "verificacion": col("Estado de verificación", i),
            "calificacion": col("Calificación SBS mostrada en video", i),
            "grupo": col("Grupo económico / respaldo", i),
            "fsd": col("Fondo de Seguro de Depósitos", i),
            "monto_minimo": col("Monto mínimo de apertura", i, True),
            "saldo_tasa_max": col("Saldo/condición para tasa máxima", i, True),
            "condicion": col("Condición clave", i),
            "abono": col("Abono de intereses", i),
            "capitalizacion": col("Capitalización", i),
            "transferencias": col("Transferencias interbancarias", i),
            "mantenimiento": col("Mantenimiento", i),
            "retiros": col("Retiros / cajeros", i),
            "alertas": col("Condiciones esenciales / alertas", i),
            "url": col("URL del producto", i),
            "video": col("Video de referencia", i),
        })
    ranking.sort(key=lambda e: -(e["trea"] or 0))

    # --- BCRP ---
    ws = wb["TASA REFERENCIA"]
    rows = rows_of(ws)
    hdr = find_row(rows, 0, "Mes")
    serie = []
    for r in rows[hdr + 1:]:
        m, v = txt(r[0]), num(r[1] if len(r) > 1 else None)
        if m and v is not None and re.match(r"\d{4}-\d{2}-\d{2}", str(m)):
            serie.append({"mes": str(m)[:10], "tasa": v})
    bcrp = {
        "vigente": num(labels_get(rows, "Tasa de referencia vigente")),
        "ultimo_mes": txt(labels_get(rows, "Último mes publicado", text=True)),
        "fuente": "https://estadisticas.bcrp.gob.pe/estadisticas/series/mensuales/resultados/PD04722MM/html",
        "serie": serie,
    }

    # --- FSD ---
    ws = wb["FSD JUN-AGO 2026"]
    rows = rows_of(ws)
    fsd = {
        "cobertura": 122000,
        "nota": txt(rows[1][0]),
        "fuente": "https://fsd.org.pe/",
        "fuente_sbs": "https://www.sbs.gob.pe/usuarios/aprende-con-la-sbs/compara-y-elige",
        "advertencia": next((txt(r[0]) for r in rows if txt(r[0]) and txt(r[0]).startswith("Importante")), None),
    }

    # --- Fichas por entidad ---
    fichas = [parse_entity(wb[s]) for s in wb.sheetnames if s not in SKIP]

    # --- Fuentes y notas ---
    ws = wb["FUENTES Y NOTAS"]
    notas = [[txt(c) for c in r] for r in rows_of(ws) if any(txt(c) for c in r)]

    data = {
        "corte": "2026-08-06",
        "ranking": ranking,
        "bcrp": bcrp,
        "fsd": fsd,
        "fichas": fichas,
        "notas": notas,
    }
    out = ROOT / "docs" / "data" / "dataset.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"{len(ranking)} entidades, {len(fichas)} fichas, {len(serie)} meses BCRP -> {out}")


def labels_get(rows, needle, text=False):
    i = find_row(rows, 0, needle)
    return rows[i][1] if i is not None and len(rows[i]) > 1 else None


if __name__ == "__main__":
    main()
