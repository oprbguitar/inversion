"""Reconstruye el Excel corrigiendo los defectos del archivo original.

Correcciones frente al original:
  * Celdas numericas guardadas con formato de fecha (S/20,000 aparecia como 1954-10-03).
  * Formulas rotas que devolvian #VALUE! al operar sobre esas fechas.
  * Ranking transpuesto (una columna por entidad) -> una fila por entidad, filtrable.
  * Sin orden, sin formato condicional y sin autofiltro.
"""
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA = json.loads((ROOT / "docs" / "data" / "dataset.json").read_text(encoding="utf-8"))
LIVE_PATH = ROOT / "docs" / "data" / "live.json"
OUT = ROOT / "docs" / "descargas" / "Ranking_Cuentas_Ahorro_Peru_Agosto_2026_MEJORADO.xlsx"

AZUL = "0B3D6B"
AZUL_CLARO = "1668AD"
BLANCO = "FFFFFF"
GRIS = "EEF2F7"
VERDE = "0F7B53"
AMBAR = "B06F00"
ROJO = "C8102E"

F_TITULO = Font(size=15, bold=True, color=AZUL)
F_SUB = Font(size=9, italic=True, color="5A6A7D")
F_CAB = Font(bold=True, color=BLANCO, size=10)
R_CAB = PatternFill("solid", fgColor=AZUL)
BORDE = Border(*[Side(style="thin", color="D8E0EA")] * 4)

# Momento exacto en que se genera el libro: se estampa en cada hoja.
GENERADO = datetime.now()
SELLO = GENERADO.strftime("%d/%m/%Y %H:%M")

MONEDA = '"S/"#,##0.00'
MONEDA0 = '"S/"#,##0'
PCT = "0.00%"


def encabezado(ws, titulo, subtitulo, ancho=8):
    """Cabecera con banda de color, titulo y sello de generacion."""
    for c in range(1, ancho + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=AZUL)
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=GRIS)

    ws["A1"] = titulo
    ws["A1"].font = Font(size=14, bold=True, color=BLANCO)
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws["A2"] = f"{subtitulo}   ·   Datos generados el {SELLO}"
    ws["A2"].font = Font(size=9, italic=True, color="3D4B5C")
    ws["A2"].alignment = Alignment(vertical="center", indent=1, wrap_text=False)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ancho)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.sheet_view.showGridLines = False


def escribir_cabecera(ws, fila, columnas):
    for c, titulo in enumerate(columnas, start=1):
        cel = ws.cell(row=fila, column=c, value=titulo)
        cel.font = F_CAB
        cel.fill = R_CAB
        cel.alignment = Alignment(vertical="center", wrap_text=True)
        cel.border = BORDE
    ws.row_dimensions[fila].height = 30


def anchos(ws, mapa):
    for col, w in mapa.items():
        ws.column_dimensions[col].width = w


def hoja_ranking(wb):
    ws = wb.create_sheet("RANKING")
    encabezado(ws, "RANKING DE CUENTAS DE AHORRO EN SOLES",
               f"Ordenado por TREA descendente. Corte: {DATA['corte']}. "
               "Cifras en formato numerico corregido. No constituye recomendacion financiera.", 14)

    cols = ["#", "Entidad", "Producto", "TREA", "Modo de apertura", "Monto minimo",
            "Saldo para tasa maxima", "Condicion clave", "Vigencia / campana",
            "Estado de verificacion", "Calificacion SBS", "Grupo economico", "FSD", "URL del producto"]
    escribir_cabecera(ws, 4, cols)

    for i, e in enumerate(DATA["ranking"], start=1):
        f = 4 + i
        vals = [i, e["entidad"], e["producto"], e["trea"], e["apertura"],
                e["monto_minimo"] or 0, e["saldo_tasa_max"] or 0, e["condicion"],
                e["vigencia"], e["verificacion"], e["calificacion"], e["grupo"], e["fsd"], e["url"]]
        for c, v in enumerate(vals, start=1):
            cel = ws.cell(row=f, column=c, value=v)
            cel.border = BORDE
            cel.alignment = Alignment(vertical="top", wrap_text=c in (8, 9))
        ws.cell(row=f, column=4).number_format = PCT
        ws.cell(row=f, column=6).number_format = MONEDA0
        ws.cell(row=f, column=7).number_format = MONEDA0
        if e["url"]:
            cel = ws.cell(row=f, column=14)
            cel.hyperlink = e["url"]
            cel.font = Font(color="1668AD", underline="single", size=9)
        ws.row_dimensions[f].height = 42

    ultima = 4 + len(DATA["ranking"])
    ws.auto_filter.ref = f"A4:N{ultima}"
    ws.freeze_panes = "C5"

    # Barras de datos dentro de la celda: la TREA se lee de un vistazo.
    ws.conditional_formatting.add(f"D5:D{ultima}", DataBarRule(
        start_type="num", start_value=0, end_type="max", color=VERDE, showValue=True))
    # Semaforo en el estado de verificacion.
    ws.conditional_formatting.add(f"J5:J{ultima}", CellIsRule(
        operator="containsText", formula=['NOT(ISERROR(SEARCH("Verificado",J5)))'],
        fill=PatternFill("solid", fgColor="DFF3E7"), font=Font(color="0F7B53", bold=True)))
    ws.conditional_formatting.add(f"J5:J{ultima}", CellIsRule(
        operator="containsText", formula=['NOT(ISERROR(SEARCH("reserva",J5)))'],
        fill=PatternFill("solid", fgColor="FBE3E6"), font=Font(color=ROJO, bold=True)))
    # Resalta las cuentas sin monto minimo.
    ws.conditional_formatting.add(f"F5:F{ultima}", CellIsRule(
        operator="equal", formula=["0"], font=Font(color=VERDE, bold=True)))
    # Filas alternas para leer tablas anchas sin perder la linea.
    ws.conditional_formatting.add(f"A5:N{ultima}", CellIsRule(
        operator="equal", formula=["MOD(ROW(),2)=0"], fill=PatternFill("solid", fgColor="F7F9FC")))

    # Grafico de barras del ranking, visible junto a la tabla.
    ch = BarChart()
    ch.type = "bar"
    ch.title = "TREA por entidad"
    ch.height, ch.width = 12, 16
    ch.legend = None
    ch.dLbls = DataLabelList()
    ch.dLbls.showVal = True
    ch.add_data(Reference(ws, min_col=4, min_row=4, max_row=ultima), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, min_row=5, max_row=ultima))
    ws.add_chart(ch, f"P4")

    anchos(ws, {"A": 4, "B": 26, "C": 30, "D": 9, "E": 22, "F": 14, "G": 18,
                "H": 46, "I": 40, "J": 26, "K": 20, "L": 26, "M": 6, "N": 40})
    return ws


def hoja_tramos(wb):
    ws = wb.create_sheet("ESCALAS DE TASA")
    encabezado(ws, "ESCALAS DE TASA POR SALDO",
               "Tramos consolidados de las 16 entidades. En el original estaban dispersos en cada hoja.", 5)
    escribir_cabecera(ws, 4, ["Entidad", "Producto", "Desde saldo", "TREA del tramo", "Descripcion del tramo"])

    f = 5
    for ficha in DATA["fichas"]:
        for t in ficha["tramos"]:
            for c, v in enumerate([ficha["entidad"], ficha["producto"], t["desde"], t["trea"], t["desc"]], start=1):
                cel = ws.cell(row=f, column=c, value=v)
                cel.border = BORDE
                cel.alignment = Alignment(vertical="top", wrap_text=c == 5)
            ws.cell(row=f, column=3).number_format = MONEDA0
            ws.cell(row=f, column=4).number_format = PCT
            f += 1

    ws.auto_filter.ref = f"A4:E{f - 1}"
    ws.freeze_panes = "A5"
    ws.conditional_formatting.add(f"D5:D{f - 1}", ColorScaleRule(
        start_type="min", start_color="FFF4E5", end_type="max", end_color="0F7B53"))
    anchos(ws, {"A": 28, "B": 32, "C": 14, "D": 14, "E": 46})


def hoja_simulador(wb):
    """Simulador con formulas vivas: el usuario cambia B5/B6 y todo recalcula."""
    ws = wb.create_sheet("SIMULADOR")
    encabezado(ws, "SIMULADOR DE INTERESES",
               "Edita el monto (B5) y el plazo (B6). Las formulas recalculan solas. "
               "TREA efectiva anual con capitalizacion mensual.", 7)

    ws["A4"] = "Parametros"; ws["A4"].font = Font(bold=True)
    ws["A5"] = "Monto depositado (S/)"; ws["B5"] = 10000; ws["B5"].number_format = MONEDA
    ws["A6"] = "Plazo (meses)"; ws["B6"] = 12
    ws["A7"] = "Aporte mensual (S/)"; ws["B7"] = 0; ws["B7"].number_format = MONEDA
    for r in (5, 6, 7):
        ws[f"B{r}"].fill = PatternFill("solid", fgColor="FFF4CC")
        ws[f"B{r}"].border = BORDE
        ws[f"A{r}"].font = Font(bold=True, size=10)

    escribir_cabecera(ws, 9, ["#", "Entidad", "TREA", "Tasa mensual equivalente",
                              "Saldo remunerado", "Interes del periodo", "Saldo final"])

    for i, e in enumerate(DATA["ranking"], start=1):
        f = 9 + i
        ficha = next((x for x in DATA["fichas"] if x["entidad"] == e["entidad"]), None)
        tope = (ficha or {}).get("simulador", {}).get("tope_remunerado")
        ws.cell(row=f, column=1, value=i)
        ws.cell(row=f, column=2, value=e["entidad"])
        ws.cell(row=f, column=3, value=e["trea"]).number_format = PCT
        ws.cell(row=f, column=4, value=f"=(1+C{f})^(1/12)-1").number_format = "0.0000%"
        # Saldo remunerado: aplica el tope de la campana cuando existe.
        remunerado = f"=MIN($B$5,{tope})" if (isinstance(tope, (int, float)) and tope > 0) else "=$B$5"
        ws.cell(row=f, column=5, value=remunerado).number_format = MONEDA
        ws.cell(row=f, column=6, value=f"=E{f}*D{f}*$B$6 + $B$7*D{f}*$B$6*($B$6-1)/2").number_format = MONEDA
        ws.cell(row=f, column=7, value=f"=$B$5 + $B$7*$B$6 + F{f}").number_format = MONEDA
        for c in range(1, 8):
            ws.cell(row=f, column=c).border = BORDE

    ultima = 9 + len(DATA["ranking"])
    ws.conditional_formatting.add(f"F10:F{ultima}", DataBarRule(
        start_type="num", start_value=0, end_type="max", color=VERDE, showValue=True))
    ws.conditional_formatting.add(f"C10:C{ultima}", ColorScaleRule(
        start_type="min", start_color="FFF4E5", end_type="max", end_color="C6E6D5"))
    ws.freeze_panes = "A10"
    anchos(ws, {"A": 4, "B": 28, "C": 10, "D": 22, "E": 18, "F": 20, "G": 18})

    ch = BarChart()
    ch.type = "bar"
    ch.title = "Interes estimado del periodo por entidad"
    ch.height, ch.width = 12, 16
    ch.legend = None
    ch.add_data(Reference(ws, min_col=6, min_row=9, max_row=ultima), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, min_row=10, max_row=ultima))
    ws.add_chart(ch, "I9")

    ws.cell(row=ultima + 2, column=1,
            value="Simulacion referencial. No incluye ITF, retenciones, comisiones no informadas, "
                  "cambios de tasa ni requisitos que el usuario no cumpla.").font = F_SUB


def hoja_bcrp(wb):
    ws = wb.create_sheet("TASA BCRP")
    b = DATA["bcrp"]
    if LIVE_PATH.exists():
        live = json.loads(LIVE_PATH.read_text(encoding="utf-8"))
        if live.get("bcrp"):
            b = {**b, **live["bcrp"]}

    encabezado(ws, "TASA DE REFERENCIA DEL BCRP",
               f"Serie mensual PD04722MM. Vigente: {b['vigente'] * 100:.2f}%. Fuente: {b['fuente']}", 4)

    escribir_cabecera(ws, 4, ["Mes", "Tasa de referencia", "Variacion (pp)"])
    serie = b.get("serie", [])
    for i, p in enumerate(serie):
        f = 5 + i
        ws.cell(row=f, column=1, value=str(p["mes"])[:7]).border = BORDE
        ws.cell(row=f, column=2, value=p["tasa"]).number_format = PCT
        ws.cell(row=f, column=2).border = BORDE
        if i:
            ws.cell(row=f, column=3, value=f"=(B{f}-B{f - 1})*100").number_format = "+0.00;-0.00;0"
        ws.cell(row=f, column=3).border = BORDE

    ultima = 4 + len(serie)
    if len(serie) > 1:
        ch = LineChart()
        ch.title = "Tasa de referencia del BCRP"
        ch.y_axis.title = "Tasa"
        ch.height, ch.width = 8, 20
        ch.add_data(Reference(ws, min_col=2, min_row=4, max_row=ultima), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=ultima))
        ws.add_chart(ch, "E4")

    ws.freeze_panes = "A5"
    anchos(ws, {"A": 12, "B": 18, "C": 16})


def hoja_fsd(wb):
    ws = wb.create_sheet("FSD Y SEGURIDAD")
    fsd = DATA["fsd"]
    encabezado(ws, "FONDO DE SEGURO DE DEPOSITOS",
               f"Cobertura vigente: S/{fsd['cobertura']:,} por depositante y por institucion. "
               f"Fuente: {fsd['fuente']}", 5)

    ws["A4"] = "Simulador de cobertura"; ws["A4"].font = Font(bold=True)
    ws["A5"] = "Tu deposito en una entidad (S/)"; ws["B5"] = 150000
    ws["A6"] = "Limite del FSD (S/)"; ws["B6"] = fsd["cobertura"]
    ws["A7"] = "Monto cubierto (S/)"; ws["B7"] = "=MIN(B5,B6)"
    ws["A8"] = "Exceso sin cobertura (S/)"; ws["B8"] = "=MAX(0,B5-B6)"
    ws["A9"] = "Entidades necesarias"; ws["B9"] = "=MAX(1,ROUNDUP(B5/B6,0))"
    for r in range(5, 10):
        ws[f"A{r}"].font = Font(bold=True, size=10)
        ws[f"B{r}"].border = BORDE
        if r != 9:
            ws[f"B{r}"].number_format = MONEDA
    ws["B5"].fill = PatternFill("solid", fgColor="FFF4CC")

    escribir_cabecera(ws, 11, ["Entidad", "Producto", "FSD", "Estado de verificacion", "Fuente oficial"])
    for i, e in enumerate(DATA["ranking"]):
        f = 12 + i
        for c, v in enumerate([e["entidad"], e["producto"], e["fsd"], e["verificacion"], e["url"]], start=1):
            cel = ws.cell(row=f, column=c, value=v)
            cel.border = BORDE
            cel.alignment = Alignment(vertical="top", wrap_text=c == 4)
        if e["url"]:
            cel = ws.cell(row=f, column=5)
            cel.hyperlink = e["url"]
            cel.font = Font(color="1668AD", underline="single", size=9)

    ws.cell(row=13 + len(DATA["ranking"]), column=1, value=fsd.get("advertencia") or "").font = F_SUB
    anchos(ws, {"A": 30, "B": 32, "C": 8, "D": 30, "E": 46})


def hoja_fichas(wb):
    ws = wb.create_sheet("FICHAS POR ENTIDAD")
    encabezado(ws, "FICHA COMPLETA POR ENTIDAD",
               "Todos los campos de las 16 hojas individuales del original, en una sola tabla filtrable.", 6)

    campos = ["Modo de apertura", "Vigencia / campana", "Condicion clave", "Abono de intereses",
              "Capitalizacion", "Mantenimiento", "Retiros / cajeros", "Transferencias", "Notas y alertas"]
    escribir_cabecera(ws, 4, ["Entidad", "Producto", "TREA"] + campos)

    origen = {"Vigencia / campana": "Vigencia / campaña", "Capitalizacion": "Capitalización",
              "Abono de intereses": "Abono de intereses", "Condicion clave": "Condición clave",
              "Retiros / cajeros": "Retiros / cajeros", "Modo de apertura": "Modo de apertura",
              "Notas y alertas": "Notas y alertas", "Mantenimiento": "Mantenimiento",
              "Transferencias": "Transferencias"}

    for i, ficha in enumerate(DATA["fichas"]):
        f = 5 + i
        c_ = ficha["campos"]
        ws.cell(row=f, column=1, value=ficha["entidad"])
        ws.cell(row=f, column=2, value=ficha["producto"])
        trea = c_.get("TREA verificada")
        try:
            ws.cell(row=f, column=3, value=float(trea)).number_format = PCT
        except (TypeError, ValueError):
            ws.cell(row=f, column=3, value=trea)
        for j, campo in enumerate(campos, start=4):
            ws.cell(row=f, column=j, value=c_.get(origen[campo]))
        for c in range(1, 4 + len(campos)):
            ws.cell(row=f, column=c).border = BORDE
            ws.cell(row=f, column=c).alignment = Alignment(vertical="top", wrap_text=c >= 4)
        ws.row_dimensions[f].height = 58

    ws.auto_filter.ref = f"A4:L{4 + len(DATA['fichas'])}"
    ws.freeze_panes = "C5"
    anchos(ws, {"A": 26, "B": 30, "C": 9, "D": 20, "E": 34, "F": 40,
                "G": 22, "H": 24, "I": 28, "J": 30, "K": 30, "L": 40})


def hoja_fuentes(wb):
    ws = wb.create_sheet("FUENTES Y NOTAS")
    encabezado(ws, "FUENTES, ALCANCES Y LIMITACIONES",
               "Registro de fuentes oficiales y campos que requieren confirmacion directa.", 6)
    filas = DATA.get("notas", [])
    inicio = next((i for i, f in enumerate(filas) if f and f[0] == "Entidad/tema"), None)
    if inicio is None:
        return
    escribir_cabecera(ws, 4, [c or "" for c in filas[inicio][:6]])
    for i, fila in enumerate(filas[inicio + 1:]):
        if not fila or not fila[0]:
            continue
        f = 5 + i
        for c, v in enumerate(fila[:6], start=1):
            cel = ws.cell(row=f, column=c, value=v)
            cel.border = BORDE
            cel.alignment = Alignment(vertical="top", wrap_text=True)
            if c == 3 and v and str(v).startswith("http"):
                cel.hyperlink = v
                cel.font = Font(color="1668AD", underline="single", size=9)
        ws.row_dimensions[f].height = 34
    ws.auto_filter.ref = f"A4:F{4 + len(filas) - inicio - 1}"
    anchos(ws, {"A": 26, "B": 26, "C": 44, "D": 34, "E": 24, "F": 50})


def hoja_portada(wb):
    ws = wb.active
    ws.title = "PORTADA"
    encabezado(ws, "RANKING DE CUENTAS DE AHORRO EN PERU — VERSION MEJORADA",
               f"Corte de datos: {DATA['corte']}", 6)

    # Panel de indicadores destacados.
    r = DATA["ranking"]
    mejor = r[0]
    promedio = sum(e["trea"] or 0 for e in r) / len(r)
    sin_min = sum(1 for e in r if not e["monto_minimo"])
    kpis = [
        ("Entidades comparadas", len(r), None),
        ("TREA mas alta", mejor["trea"], PCT),
        ("TREA promedio", promedio, PCT),
        ("Sin monto minimo", sin_min, None),
        ("Tasa de referencia BCRP", DATA["bcrp"]["vigente"], PCT),
        ("Cobertura FSD", DATA["fsd"]["cobertura"], MONEDA0),
    ]
    for i, (etiqueta, valor, fmt) in enumerate(kpis):
        col = 1 + (i % 3) * 2
        fila = 4 + (i // 3) * 3
        et = ws.cell(row=fila, column=col, value=etiqueta.upper())
        et.font = Font(size=8, bold=True, color="5A6A7D")
        val = ws.cell(row=fila + 1, column=col, value=valor)
        val.font = Font(size=18, bold=True, color=AZUL)
        if fmt:
            val.number_format = fmt
        for c in (col, col + 1):
            for f in (fila, fila + 1):
                ws.cell(row=f, column=c).fill = PatternFill("solid", fgColor="F2F6FA")
                ws.cell(row=f, column=c).border = Border(
                    left=Side(style="thick", color=AZUL_CLARO) if c == col else None,
                    top=Side(style="thin", color="D8E0EA"),
                    bottom=Side(style="thin", color="D8E0EA"),
                )
    ws.cell(row=9, column=1,
            value=f"Libro generado el {SELLO}. La fecha de tu descarga va en el nombre del archivo."
            ).font = Font(size=10, bold=True, color=AZUL_CLARO)

    lineas = [
        ("", ""),
        ("Contenido del libro", ""),
        ("RANKING", "16 entidades ordenadas por TREA, con filtros y formato condicional."),
        ("ESCALAS DE TASA", "Tramos de tasa por saldo de todas las entidades, consolidados."),
        ("SIMULADOR", "Formulas vivas: cambia el monto y el plazo y recalcula las 16 entidades."),
        ("TASA BCRP", "Serie mensual de la tasa de referencia, con grafico."),
        ("FSD Y SEGURIDAD", "Cobertura del Fondo de Seguro de Depositos y simulador de cobertura."),
        ("FICHAS POR ENTIDAD", "Todos los campos de las 16 hojas individuales, en una tabla filtrable."),
        ("FUENTES Y NOTAS", "Fuentes oficiales y nivel de confianza de cada dato."),
        ("", ""),
        ("Correcciones aplicadas al archivo original", ""),
        ("1", "Celdas numericas guardadas con formato de fecha (S/20,000 se veia como 1954-10-03)."),
        ("2", "Formulas que devolvian #VALUE! por operar sobre esas fechas."),
        ("3", "Ranking transpuesto (una columna por entidad) convertido a una fila por entidad."),
        ("4", "Ranking sin ordenar: ahora va por TREA descendente y numerado."),
        ("5", "Sin autofiltro ni formato condicional: se anadieron en todas las tablas."),
        ("6", "Escalas de tasa dispersas en 16 hojas: consolidadas en una sola."),
        ("", ""),
        ("Aviso", "Documento informativo. No es asesoria financiera ni recomendacion de contratacion. "
                  "Verifique el contrato, la cartilla y el tarifario vigentes antes de abrir una cuenta."),
    ]
    for i, (a, b) in enumerate(lineas, start=11):
        ca = ws.cell(row=i, column=1, value=a)
        ca.font = Font(bold=True, size=10, color=AZUL)
        ca.alignment = Alignment(vertical="top")
        cb = ws.cell(row=i, column=2, value=b)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        # Los encabezados de seccion no llevan segunda columna: se realzan.
        if a and not b:
            ca.font = Font(bold=True, size=11, color=BLANCO)
            for c in range(1, 7):
                ws.cell(row=i, column=c).fill = PatternFill("solid", fgColor=AZUL_CLARO)
    anchos(ws, {"A": 28, "B": 44, "C": 16, "D": 16, "E": 16, "F": 16})
    ws.column_dimensions["B"].width = 92


def main():
    wb = Workbook()
    hoja_portada(wb)
    hoja_ranking(wb)
    hoja_tramos(wb)
    hoja_simulador(wb)
    hoja_bcrp(wb)
    hoja_fsd(wb)
    hoja_fichas(wb)
    hoja_fuentes(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"{len(wb.sheetnames)} hojas -> {OUT} ({OUT.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
